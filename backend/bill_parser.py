import io
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
from drive_client import list_folders, list_spreadsheets, download_xlsx

BILL_FOLDER_ID = "1IN5qeH4vC2Z4FabEdjb6U1PRKF0Em7GI"

DELAY_BUCKETS = ["T0", "T1", "T2", "T3", "T4", "T5", "T5Plus"]
# Column positions confirmed from spreadsheet inspection:
# T0=11, T1=15, T2=19, T3=23, T4=27, T5=31, T5Plus=35
# Each bucket: bills(+0), % bills(+1), amount(+2), % amount(+3)
BUCKET_BASE_COLS = [11, 15, 19, 23, 27, 31, 35]


def _safe_num(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.replace('%', '').strip()
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _safe_int(v):
    n = _safe_num(v)
    return int(n) if n is not None else None


def _clean_pao_name(raw: str):
    """Strip numeric code prefix from PAO name. Returns (clean_name, code)."""
    raw = raw.strip()
    m = re.match(r'^(\d+)[-–]\s*', raw)
    if m:
        return raw[m.end():].strip(), m.group(1)
    return raw, None


def _find_period(rows):
    """Scan early rows for a Period: label and return the value in the same row."""
    for row in rows[:10]:
        for i, cell in enumerate(row):
            if cell and "period" in str(cell).lower():
                for v in row[i+1:]:
                    if v and "period" not in str(v).lower():
                        return str(v).strip()
    return None


def _parse_ebm(wb: openpyxl.Workbook) -> dict:
    """EBM-01: Controller wise E-bill count and amount percentage report."""
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    period = _find_period(rows)
    header_row_idx = None

    for i, row in enumerate(rows):
        if any(c and "controller code" in str(c).lower() for c in row):
            header_row_idx = i
            break

    if header_row_idx is None:
        return {"period": period, "paos": []}

    paos = []
    for row in rows[header_row_idx + 1:]:
        if not row or row[0] is None:
            continue
        raw = str(row[0]).strip()
        if not raw or "total" in raw.lower():
            continue

        pao_raw = row[1] if len(row) > 1 else None
        if pao_raw is None:
            continue
        pao_name, pao_code = _clean_pao_name(str(pao_raw))

        paos.append({
            "pao_code": pao_code,
            "pao_name": pao_name,
            "total_bills":     _safe_int(row[3]) if len(row) > 3  else None,
            "total_amount":    _safe_num(row[4]) if len(row) > 4  else None,
            "normal_bills":    _safe_int(row[5]) if len(row) > 5  else None,
            "normal_amount":   _safe_num(row[6]) if len(row) > 6  else None,
            "ebill_count":     _safe_int(row[7]) if len(row) > 7  else None,
            "ebill_amount":    _safe_num(row[8]) if len(row) > 8  else None,
            "pct_ebill_count":  _safe_num(row[9])  if len(row) > 9  else None,
            "pct_ebill_amount": _safe_num(row[10]) if len(row) > 10 else None,
        })

    return {"period": period, "paos": paos}


def _parse_delay(wb: openpyxl.Workbook, sanction_type: str) -> dict:
    """TM-02: Ministry wise PAO delay summary (normal or ebill).

    Spreadsheet layout (confirmed):
      Row N  : header  — col[0]='Ministry', col[1]='PAO', col[4]=Token, col[5]=Closed, ...
      Row N+1: sub-hdr — col labels A/B/B1/C/D/E and T0..T5Plus
      Row N+2: 'Grand Total' row (col[1]='Grand Total:') — skip
      Row N+3: ministry subtotal (col[0]='030-Ministry...', col[1]='Ministry wise Total:') — skip
      Rows N+4+: individual PAO rows — col[0]=None, col[1]='code-PAO name'
    """
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    period = _find_period(rows)
    header_row_idx = None

    for i, row in enumerate(rows):
        if any(c and str(c).strip() == "Ministry" for c in row):
            header_row_idx = i
            break

    if header_row_idx is None:
        return {"period": period, "sanction_type": sanction_type, "paos": []}

    def g(row, idx):
        return _safe_num(row[idx]) if len(row) > idx else None

    paos = []
    for row in rows[header_row_idx + 2:]:
        if not row:
            continue

        pao_raw = row[1] if len(row) > 1 else None
        if pao_raw is None:
            continue
        pao_str = str(pao_raw).strip()
        if not pao_str or "total" in pao_str.lower():
            continue

        pao_name, pao_code = _clean_pao_name(pao_str)

        bucket_data = {}
        for bucket, base in zip(DELAY_BUCKETS, BUCKET_BASE_COLS):
            bucket_data[f"{bucket}_bills"]  = g(row, base)
            bucket_data[f"{bucket}_pct"]    = g(row, base + 1)
            bucket_data[f"{bucket}_amount"] = g(row, base + 2)

        paos.append({
            "pao_code": pao_code,
            "pao":      pao_name,
            "total_bills_token": g(row, 4),
            "closed":    g(row, 5),
            "cancelled": g(row, 7),
            "returned":  g(row, 8),
            "pending":   g(row, 10),
            **bucket_data,
        })

    return {"period": period, "sanction_type": sanction_type, "paos": paos}


def _process_week(wf):
    files = list_spreadsheets(wf["id"])
    if not files:
        return None

    week_entry = {
        "period":       wf["name"],
        "ebm":          None,
        "delay_normal": None,
        "delay_ebill":  None,
    }

    def _fetch_and_parse(xf):
        name_lower = xf["name"].lower()
        if "pc-04" in name_lower or "e-payment" in name_lower or "epayment" in name_lower:
            return None
        try:
            content = download_xlsx(xf["id"], xf["mimeType"])
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        except Exception:
            return None
        if "ebm" in name_lower or "ebill count" in name_lower:
            return ("ebm", _parse_ebm(wb))
        elif "tm-02" in name_lower or "pao delay" in name_lower:
            key = "delay_ebill" if "ebill" in name_lower else "delay_normal"
            return (key, _parse_delay(wb, "ebill" if "ebill" in name_lower else "normal"))
        return None

    with ThreadPoolExecutor(max_workers=3) as pool:
        for result in pool.map(_fetch_and_parse, files):
            if result:
                week_entry[result[0]] = result[1]

    return week_entry


def parse_bill() -> dict:
    """Parse all weeks from Bill Monitoring Drive folder.

    Files used per week:
      - EBM-01  Controller Wise Ebill count and Amount Percentage Report
      - TM-02   Ministry wise PAO Delay Summary_normal
      - TM-02   Ministry wise PAO Delay Summary_ebill

    PC-04 is intentionally excluded.
    """
    month_folders = list_folders(BILL_FOLDER_ID)

    week_folders_all = []
    for mf in sorted(month_folders, key=lambda f: f["name"]):
        for wf in sorted(list_folders(mf["id"]), key=lambda f: f["name"]):
            week_folders_all.append(wf)

    all_weeks = []
    with ThreadPoolExecutor(max_workers=4) as pool:
        futures = {pool.submit(_process_week, wf): wf["name"] for wf in week_folders_all}
        results = {}
        for future in as_completed(futures):
            name = futures[future]
            entry = future.result()
            if entry:
                results[name] = entry

    for wf in week_folders_all:
        if wf["name"] in results:
            all_weeks.append(results[wf["name"]])

    return {"weeks": all_weeks}
