import io
import openpyxl
from drive_client import list_folders, list_spreadsheets, download_xlsx

IAW_FOLDER_ID = "1x9KRZLOmQxVPEczovppv2bYq1fkqeh0t"
ZONES = ["NZ", "SZ", "WZ", "EZ"]
ZONE_LABELS = {"NZ": "North Zone", "SZ": "South Zone", "WZ": "West Zone", "EZ": "East Zone"}


def _safe_int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _parse_zone_sheet(ws) -> dict:
    """
    Returns zone-level totals (from the Total row) and per-office rows.
    Zone totals use the spreadsheet's own Total row as the authoritative number.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"opening": None, "raised": None, "settled": None, "closing": None, "offices": []}

    zone_totals = {"opening": None, "raised": None, "settled": None, "closing": None}
    offices = []

    for row in rows[1:]:
        if not row or row[0] is None or str(row[0]).strip() == "":
            continue
        label = str(row[0]).strip()
        o = _safe_int(row[1]) if len(row) > 1 else None
        r = _safe_int(row[2]) if len(row) > 2 else None
        s = _safe_int(row[3]) if len(row) > 3 else None
        c = _safe_int(row[4]) if len(row) > 4 else None

        if label == "Total":
            zone_totals = {"opening": o, "raised": r, "settled": s, "closing": c}
        elif label != "Name of Office":
            offices.append({"office": label, "opening": o, "raised": r, "settled": s, "closing": c})

    return {**zone_totals, "offices": offices}


def parse_iaw() -> dict:
    # Find "Paras Status" subfolder
    top_folders = list_folders(IAW_FOLDER_ID)
    paras_status = next(
        (f for f in top_folders if "paras status" in f["name"].lower()),
        None,
    )
    search_root = paras_status["id"] if paras_status else IAW_FOLDER_ID

    month_folders = list_folders(search_root)
    months = []

    for mf in sorted(month_folders, key=lambda f: f["name"]):
        month_name = mf["name"].replace(",", " ").strip()
        files = list_spreadsheets(mf["id"])
        if not files:
            continue

        xf = files[0]
        content = download_xlsx(xf["id"], xf["mimeType"])
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        zones = []
        for zone in ZONES:
            if zone not in wb.sheetnames:
                continue
            parsed = _parse_zone_sheet(wb[zone])
            zones.append({
                "zone": zone,
                "zone_label": ZONE_LABELS[zone],
                "opening": parsed["opening"],
                "raised": parsed["raised"],
                "settled": parsed["settled"],
                "closing": parsed["closing"],
                "offices": parsed["offices"],
            })

        months.append({"month": month_name, "zones": zones})

    return {"months": months}
