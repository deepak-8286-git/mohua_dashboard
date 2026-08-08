import io
import re
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl

from drive_client import get_service, list_folders, download_xlsx

GEM_FOLDER_ID    = "1Z4DEbRGqN6gJudyPhR8dro7gl2oOUpQe"
DUES_FOLDER_ID   = "1dLTcB4KCI5UKW0dmT7UlWFSGuq9C_EuB"
AGEWISE_FOLDER_ID = "1Y4jLxA-rqKMCySG2fi646jlv2ff1sgwM"


def _safe(v):
    try:
        f = float(v)
        return round(f, 6) if f else 0.0
    except (TypeError, ValueError):
        return 0.0


def _is_grand_total(name):
    return name and "grand total" in str(name).lower()


def _parse_dues(wb):
    """Parse Sheet1 (org summary) and Sheet2 (office drill-down)."""
    # ── Sheet1: org-level ─────────────────────────────────────────────
    ws1 = wb["Sheet1"]
    rows1 = [r for r in ws1.iter_rows(values_only=True) if any(v is not None for v in r)]

    orgs = []
    for row in rows1[3:]:  # skip title, unit, header
        name = str(row[0]).strip() if row[0] else ""
        if not name or _is_grand_total(name):
            continue
        orgs.append({
            "org":   name,
            "mse":   _safe(row[1]),
            "others": _safe(row[2]),
            "total": _safe(row[3]),
        })

    # ── Sheet2: office drill-down ──────────────────────────────────────
    ws2 = wb["Sheet2"] if "Sheet2" in wb.sheetnames else None
    offices_by_org = {}

    if ws2:
        rows2 = [r for r in ws2.iter_rows(values_only=True) if any(v is not None for v in r)]
        current_org = None
        for row in rows2[3:]:
            name = str(row[0]).strip() if row[0] else ""
            if not name or _is_grand_total(name):
                continue
            has_amounts = any(row[i] not in (None, 0, 0.0) for i in [1, 2, 3])
            if not has_amounts and len([v for v in row if v is not None]) == 1:
                current_org = name
                if current_org not in offices_by_org:
                    offices_by_org[current_org] = []
            elif current_org:
                offices_by_org[current_org].append({
                    "office": name,
                    "mse":    _safe(row[1]),
                    "others": _safe(row[2]),
                    "total":  _safe(row[3]),
                })

    return {"orgs": orgs, "offices_by_org": offices_by_org}


def _parse_agewise(wb):
    """Parse Agewise sheet — buckets vary by month, store whatever is present."""
    ws = wb.active
    rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]

    # Row 2 (index 2) is the header
    header = [str(v).strip() if v else "" for v in rows[2]]
    # cols 0-3: Org, MSE, OTHERS, Total; cols 4+: age buckets
    age_bucket_names = header[4:]

    data = []
    for row in rows[3:]:
        name = str(row[0]).strip() if row[0] else ""
        if not name or _is_grand_total(name):
            continue
        buckets = {age_bucket_names[i]: _safe(row[4 + i])
                   for i in range(len(age_bucket_names))
                   if 4 + i < len(row)}
        data.append({
            "org":     name,
            "mse":     _safe(row[1]),
            "others":  _safe(row[2]),
            "total":   _safe(row[3]),
            "buckets": buckets,
        })

    return {"bucket_names": age_bucket_names, "orgs": data}


def _month_sort_key(name):
    months = ["january","february","march","april","may","june",
              "july","august","september","october","november","december"]
    low = name.lower()
    for i, m in enumerate(months):
        if m in low:
            year = re.search(r'\d{4}', name)
            return (int(year.group()) if year else 0, i)
    return (0, 0)


def _list_xlsx_in_folder(folder_id):
    svc = get_service()
    res = svc.files().list(
        q=f"'{folder_id}' in parents and trashed=false",
        fields="files(id,name,mimeType)"
    ).execute()
    return res.get("files", [])


def _fetch_and_parse(file_info, parser):
    try:
        content = download_xlsx(file_info["id"], file_info["mimeType"])
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        return parser(wb)
    except Exception as e:
        return None


def parse_gem() -> dict:
    dues_files   = sorted(_list_xlsx_in_folder(DUES_FOLDER_ID),   key=lambda f: _month_sort_key(f["name"]))
    agewise_files = sorted(_list_xlsx_in_folder(AGEWISE_FOLDER_ID), key=lambda f: _month_sort_key(f["name"]))

    def _extract_month_label(name):
        # Normalize any filename to "Month YYYY"
        # "June 2026.xlsx" → "June 2026"
        # "Age wise pendency in June 2026.xlsx" → "June 2026"
        m = re.search(
            r'(january|february|march|april|may|june|july|august|'
            r'september|october|november|december)\s+\d{4}',
            name, re.IGNORECASE
        )
        if m:
            parts = m.group().split()
            return f"{parts[0].title()} {parts[1]}"
        return re.sub(r'\.(xlsx?|xls)$', '', name, flags=re.IGNORECASE).strip()

    months = []
    with ThreadPoolExecutor(max_workers=4) as pool:
        dues_futures   = {pool.submit(_fetch_and_parse, f, _parse_dues):   _extract_month_label(f["name"]) for f in dues_files}
        agewise_futures = {pool.submit(_fetch_and_parse, f, _parse_agewise): _extract_month_label(f["name"]) for f in agewise_files}

        dues_results   = {}
        agewise_results = {}

        for fut, label in dues_futures.items():
            r = fut.result()
            if r: dues_results[label] = r

        for fut, label in agewise_futures.items():
            r = fut.result()
            if r: agewise_results[label] = r

    all_labels = sorted(set(dues_results) | set(agewise_results), key=_month_sort_key)
    for label in all_labels:
        months.append({
            "month":   label,
            "dues":    dues_results.get(label),
            "agewise": agewise_results.get(label),
        })

    return {"months": months}
