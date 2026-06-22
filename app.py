import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import openpyxl
import io
import warnings
from google.oauth2 import service_account
from googleapiclient.discovery import build

warnings.filterwarnings("ignore")

st.set_page_config(
    page_title="IAW · Audit Paras Monitor",
    page_icon="🏛",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Constants ──────────────────────────────────────────────────────────────────
CREDENTIALS_FILE = "credentials/cosmic-octane-499906-q0-88ac81084d22.json"
IAW_FOLDER_ID = "1x9KRZLOmQxVPEczovppv2bYq1fkqeh0t"
SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
GSHEET_MIME = "application/vnd.google-apps.spreadsheet"

ZONES = ["NZ", "SZ", "WZ", "EZ"]
ZONE_LABELS = {"NZ": "North Zone", "SZ": "South Zone", "WZ": "West Zone", "EZ": "East Zone"}
ZONE_COLORS = {"NZ": "#4F9CF9", "SZ": "#E8813A", "WZ": "#38B089", "EZ": "#D94F3D"}

# ── CSS injection ──────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Rajdhani:wght@400;500;600;700&family=Inter:wght@300;400;500;600&family=JetBrains+Mono:wght@400;600&display=swap');

html, body, [data-testid="stApp"] {
    background-color: #0A1628 !important;
    color: #E8EDF5 !important;
}
.main { background-color: #0A1628 !important; }
.block-container {
    padding-top: 1.5rem !important;
    padding-left: 2.5rem !important;
    padding-right: 2.5rem !important;
    max-width: 1440px !important;
}

#MainMenu, footer, [data-testid="stHeader"], [data-testid="stToolbar"],
[data-testid="stDecoration"] { display: none !important; }

[data-testid="stSidebar"] {
    background-color: #0D1B2E !important;
    border-right: 1px solid #1A2A40 !important;
}
[data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
    gap: 0 !important;
}
[data-testid="stSidebar"] label {
    font-family: 'Inter', sans-serif !important;
    font-size: 0.68rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.09em !important;
    text-transform: uppercase !important;
    color: #5A7090 !important;
}
[data-testid="stSidebar"] .stSelectbox > div > div,
[data-testid="stSidebar"] .stMultiSelect > div > div {
    background-color: #142030 !important;
    border-color: #1A2A40 !important;
    color: #E8EDF5 !important;
    font-family: 'Inter', sans-serif !important;
    font-size: 0.8rem !important;
}

hr { border-color: #1A2A40 !important; }

.stButton > button {
    background-color: transparent !important;
    border: 1px solid #1A2A40 !important;
    color: #7A8FA8 !important;
    font-family: 'Inter', sans-serif !important;
    font-size: 0.72rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.06em !important;
    border-radius: 3px !important;
    transition: border-color 0.15s, color 0.15s !important;
    width: 100% !important;
}
.stButton > button:hover {
    border-color: #E8813A !important;
    color: #E8813A !important;
    background-color: transparent !important;
}

.stDataFrame { border-radius: 4px !important; overflow: hidden !important; }
[data-testid="stSpinner"] p { color: #E8813A !important; }

div[data-testid="stMetric"] { display: none !important; }
</style>
""", unsafe_allow_html=True)


# ── Drive helpers ──────────────────────────────────────────────────────────────
@st.cache_resource
def get_drive_service():
    creds = service_account.Credentials.from_service_account_file(
        CREDENTIALS_FILE, scopes=SCOPES
    )
    return build("drive", "v3", credentials=creds)


@st.cache_data(ttl=300)
def load_all_data():
    service = get_drive_service()

    top_folders = service.files().list(
        q=f"'{IAW_FOLDER_ID}' in parents and mimeType='application/vnd.google-apps.folder'",
        fields="files(id, name)",
    ).execute().get("files", [])

    paras_status_id = next(
        (f["id"] for f in top_folders if "paras status" in f["name"].lower()),
        IAW_FOLDER_ID,
    )

    month_folders = service.files().list(
        q=f"'{paras_status_id}' in parents and mimeType='application/vnd.google-apps.folder'",
        fields="files(id, name)",
    ).execute().get("files", [])

    records = []
    for folder in month_folders:
        month_name, folder_id = folder["name"], folder["id"]
        files = service.files().list(
            q=f"'{folder_id}' in parents and (mimeType='{XLSX_MIME}' or mimeType='{GSHEET_MIME}')",
            fields="files(id, name, mimeType)",
        ).execute().get("files", [])

        for xf in files:
            if xf["mimeType"] == GSHEET_MIME:
                content = service.files().export_media(fileId=xf["id"], mimeType=XLSX_MIME).execute()
            else:
                content = service.files().get_media(fileId=xf["id"]).execute()

            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            for zone in ZONES:
                if zone not in wb.sheetnames:
                    continue
                rows = list(wb[zone].iter_rows(values_only=True))

                def safe(v):
                    try:
                        return int(v)
                    except (TypeError, ValueError):
                        return None

                for row in rows[1:]:
                    if row[0] in (None, "Total", "Name of Office") or not str(row[0]).strip():
                        continue
                    records.append({
                        "month": month_name.replace(",", " "),
                        "zone": zone,
                        "zone_label": ZONE_LABELS[zone],
                        "office": row[0],
                        "opening": safe(row[1]),
                        "raised": safe(row[2]),
                        "settled": safe(row[3]),
                        "closing": safe(row[4]),
                    })

    df = pd.DataFrame(records)
    df["month_dt"] = pd.to_datetime(df["month"], format="%B %Y")
    return df.sort_values("month_dt")


# ── Chart layout factory ───────────────────────────────────────────────────────
def chart_layout(**overrides):
    base = dict(
        paper_bgcolor="#111C2D",
        plot_bgcolor="#111C2D",
        font=dict(family="Inter, sans-serif", color="#7A8FA8", size=11),
        xaxis=dict(gridcolor="#1A2A40", zerolinecolor="#1A2A40",
                   tickfont=dict(color="#7A8FA8", size=11), linecolor="#1A2A40"),
        yaxis=dict(gridcolor="#1A2A40", zerolinecolor="#1A2A40",
                   tickfont=dict(color="#7A8FA8", size=11), linecolor="#1A2A40"),
        legend=dict(
            bgcolor="rgba(0,0,0,0)", borderwidth=0,
            font=dict(size=11, color="#7A8FA8"),
            orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
        ),
        margin=dict(l=4, r=4, t=24, b=4),
        height=360,
        hoverlabel=dict(bgcolor="#162033", bordercolor="#1A2A40",
                        font=dict(family="Inter", size=12, color="#E8EDF5")),
    )
    base.update(overrides)
    return base


# ── HTML helpers ───────────────────────────────────────────────────────────────
def divider():
    st.markdown('<div style="height:1px;background:#1A2A40;margin:1rem 0;"></div>',
                unsafe_allow_html=True)


def section_heading(title):
    st.markdown(
        f'<div style="font-family:Rajdhani,sans-serif;font-size:0.8rem;font-weight:600;'
        f'letter-spacing:0.12em;text-transform:uppercase;color:#5A7090;'
        f'display:flex;align-items:center;gap:10px;margin-bottom:0.75rem;margin-top:0.25rem;">'
        f'{title}'
        f'<div style="flex:1;height:1px;background:#1A2A40;"></div>'
        f'</div>',
        unsafe_allow_html=True,
    )


# ── Load data ──────────────────────────────────────────────────────────────────
with st.spinner("Loading from Google Drive…"):
    df = load_all_data()

month_labels = [pd.Timestamp(m).strftime("%B %Y") for m in df["month_dt"].sort_values().unique()]
latest_month = month_labels[-1]
prev_month = month_labels[-2] if len(month_labels) > 1 else None
df_prev_all = df[df["month"] == prev_month] if prev_month else None


# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="padding:1.4rem 1rem 0.8rem;border-bottom:1px solid #1A2A40;margin-bottom:1rem;">
        <div style="font-family:Inter,sans-serif;font-size:0.6rem;font-weight:600;
                    letter-spacing:0.15em;text-transform:uppercase;color:#5A7090;margin-bottom:6px;">
            Ministry of Housing &amp; Urban Affairs
        </div>
        <div style="font-family:Rajdhani,sans-serif;font-size:1.55rem;font-weight:700;
                    color:#E8EDF5;line-height:1.1;letter-spacing:0.02em;">
            Internal Audit Wing
        </div>
        <div style="font-family:JetBrains Mono,monospace;font-size:0.62rem;
                    color:#E8813A;margin-top:5px;letter-spacing:0.05em;">
            PARAS STATUS MONITOR
        </div>
    </div>
    """, unsafe_allow_html=True)

    selected_month = st.selectbox("Reporting Month", options=month_labels[::-1], index=0)
    st.markdown("<div style='height:0.5rem;'></div>", unsafe_allow_html=True)
    selected_zones = st.multiselect(
        "Zones", options=ZONES, default=ZONES,
        format_func=lambda z: f"{z} — {ZONE_LABELS[z]}",
    )

    # Zone snapshot cards
    df_sidebar = df[df["month"] == selected_month].dropna(subset=["closing"])
    if not df_sidebar.empty and selected_zones:
        st.markdown("""
        <div style="font-family:Inter,sans-serif;font-size:0.6rem;font-weight:600;
                    letter-spacing:0.12em;text-transform:uppercase;color:#5A7090;
                    margin:1.2rem 0 0.5rem;">
            Zone Snapshot
        </div>
        """, unsafe_allow_html=True)
        for zone in selected_zones:
            zd = df_sidebar[df_sidebar["zone"] == zone]
            if zd.empty:
                continue
            closing = int(zd["closing"].sum())
            settled = int(zd["settled"].sum())
            raised = int(zd["raised"].sum())
            color = ZONE_COLORS.get(zone, "#888")
            st.markdown(
                f'<div style="background:#12202F;border-left:3px solid {color};border-radius:3px;'
                f'padding:9px 12px;margin-bottom:5px;display:flex;justify-content:space-between;align-items:center;">'
                f'<div>'
                f'<div style="font-family:Rajdhani,sans-serif;font-size:0.78rem;font-weight:700;'
                f'letter-spacing:0.06em;color:{color};">{zone} &nbsp;·&nbsp; {ZONE_LABELS[zone]}</div>'
                f'<div style="font-family:Inter,sans-serif;font-size:0.62rem;color:#5A7090;margin-top:2px;">'
                f'{settled} settled &nbsp;·&nbsp; {raised} raised</div>'
                f'</div>'
                f'<div style="font-family:Rajdhani,sans-serif;font-size:1.5rem;font-weight:700;color:#E8EDF5;">'
                f'{closing:,}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )

    st.markdown("<div style='height:1rem;'></div>", unsafe_allow_html=True)
    if st.button("↻  Refresh Data"):
        st.cache_data.clear()
        st.rerun()
    st.markdown(
        '<div style="font-family:Inter,sans-serif;font-size:0.6rem;color:#3A4E63;'
        'text-align:center;margin-top:6px;">Auto-refreshes every 5 minutes</div>',
        unsafe_allow_html=True,
    )


# ── Filtered data ──────────────────────────────────────────────────────────────
df_sel = df[(df["month"] == selected_month) & (df["zone"].isin(selected_zones))]
df_num = df_sel.dropna(subset=["closing"])

# ── Page header ────────────────────────────────────────────────────────────────
zone_badges = " &nbsp;·&nbsp; ".join(
    f'<span style="color:{ZONE_COLORS.get(z,"#888")}">{ZONE_LABELS[z]}</span>'
    for z in selected_zones
)
st.markdown(
    f'<div style="display:flex;justify-content:space-between;align-items:flex-end;margin-bottom:1.2rem;">'
    f'<div>'
    f'<div style="font-family:Inter,sans-serif;font-size:0.6rem;font-weight:600;'
    f'letter-spacing:0.16em;text-transform:uppercase;color:#E8813A;margin-bottom:5px;">'
    f'Internal Audit Wing &nbsp;·&nbsp; Audit Paras Status</div>'
    f'<div style="font-family:Rajdhani,sans-serif;font-size:2.4rem;font-weight:700;'
    f'color:#E8EDF5;line-height:1;letter-spacing:0.02em;">{selected_month}</div>'
    f'<div style="font-family:Inter,sans-serif;font-size:0.72rem;color:#5A7090;margin-top:5px;">'
    f'{zone_badges}</div>'
    f'</div>'
    f'<div style="text-align:right;">'
    f'<div style="font-family:JetBrains Mono,monospace;font-size:0.6rem;color:#3A4E63;">MoHUA / IAW</div>'
    f'<div style="font-family:JetBrains Mono,monospace;font-size:0.6rem;color:#3A4E63;margin-top:2px;">'
    f'PARAS MONITOR v1.0</div>'
    f'</div>'
    f'</div>'
    f'<div style="height:1px;background:linear-gradient(to right,#E8813A 0%,#1A2A40 60%);'
    f'margin-bottom:1.6rem;"></div>',
    unsafe_allow_html=True,
)

# ── KPI cards ──────────────────────────────────────────────────────────────────
total_opening = int(df_num["opening"].sum())
total_raised = int(df_num["raised"].sum())
total_settled = int(df_num["settled"].sum())
total_closing = int(df_num["closing"].sum())
settlement_rate = (
    (total_settled / (total_opening + total_raised) * 100)
    if (total_opening + total_raised) > 0 else 0
)

prev_closing_val = (
    df_prev_all[df_prev_all["zone"].isin(selected_zones)]
    .dropna(subset=["closing"])["closing"].sum()
    if df_prev_all is not None else None
)
delta_closing = int(total_closing - prev_closing_val) if prev_closing_val else None
delta_color = "#D94F3D" if (delta_closing or 0) > 0 else "#38B089"
delta_arrow = "▲" if (delta_closing or 0) > 0 else "▼"

def kpi_card(label, value, sub, accent, extra_html=""):
    return (
        f'<div style="background:#12202F;border-top:2px solid {accent};border-radius:4px;'
        f'padding:18px 20px 15px;height:100%;box-sizing:border-box;">'
        f'<div style="font-family:Inter,sans-serif;font-size:0.62rem;font-weight:600;'
        f'letter-spacing:0.12em;text-transform:uppercase;color:#5A7090;">{label}</div>'
        f'<div style="font-family:Rajdhani,sans-serif;font-size:2.5rem;font-weight:700;'
        f'color:{accent};line-height:1.05;margin:6px 0 3px;">{value}</div>'
        f'<div style="font-family:Inter,sans-serif;font-size:0.65rem;color:#3A4E63;">{sub}</div>'
        f'{extra_html}'
        f'</div>'
    )

delta_extra = (
    f'<div style="font-family:Inter,sans-serif;font-size:0.68rem;color:{delta_color};margin-top:6px;">'
    f'{delta_arrow} {abs(delta_closing):,} vs {prev_month}</div>'
    if delta_closing is not None else ""
)

rate_w = f"{min(settlement_rate, 100):.1f}%"
rate_color = "#38B089" if settlement_rate >= 5 else "#E8813A"
rate_extra = (
    f'<div style="margin-top:10px;background:#1A2A40;border-radius:2px;height:3px;">'
    f'<div style="width:{rate_w};height:3px;background:{rate_color};border-radius:2px;'
    f'transition:width 0.4s ease;"></div></div>'
)

c1, c2, c3, c4, c5 = st.columns(5)
with c1:
    st.markdown(kpi_card("Opening Balance", f"{total_opening:,}",
                          "Paras at start of month", "#4F9CF9"), unsafe_allow_html=True)
with c2:
    st.markdown(kpi_card("Paras Raised", f"{total_raised:,}",
                          "New paras added this month", "#E8813A"), unsafe_allow_html=True)
with c3:
    st.markdown(kpi_card("Paras Settled", f"{total_settled:,}",
                          "Paras resolved this month", "#38B089"), unsafe_allow_html=True)
with c4:
    st.markdown(kpi_card("Closing Balance", f"{total_closing:,}",
                          "Paras outstanding at month-end", "#E8EDF5", delta_extra),
                unsafe_allow_html=True)
with c5:
    st.markdown(kpi_card("Settlement Rate", f"{settlement_rate:.1f}%",
                          "Resolved vs total exposure", rate_color, rate_extra),
                unsafe_allow_html=True)

st.markdown("<div style='height:1.8rem;'></div>", unsafe_allow_html=True)

# ── Charts row 1 ───────────────────────────────────────────────────────────────
col_a, col_b = st.columns([3, 2])

with col_a:
    section_heading("Outstanding Paras by Zone")
    zone_sum = (
        df_num.groupby("zone")[["opening", "raised", "settled", "closing"]]
        .sum().reset_index()
    )
    zone_sum["zone_label"] = zone_sum["zone"].map(ZONE_LABELS)

    fig_zone = go.Figure()
    bars = [("Opening", "#2B4069"), ("Raised", "#E8813A"),
            ("Settled", "#38B089"), ("Closing", "#4F9CF9")]
    for name, color in bars:
        fig_zone.add_trace(go.Bar(
            name=name,
            x=zone_sum["zone_label"],
            y=zone_sum[name.lower()],
            marker=dict(color=color, line=dict(width=0)),
        ))
    fig_zone.update_layout(**chart_layout(barmode="group", yaxis_title="No. of Paras"))
    st.plotly_chart(fig_zone, use_container_width=True, config={"displayModeBar": False})

with col_b:
    section_heading("Zone Share — Closing Balance")
    if not zone_sum.empty:
        fig_donut = go.Figure(go.Pie(
            labels=zone_sum["zone_label"],
            values=zone_sum["closing"],
            hole=0.62,
            marker=dict(
                colors=[ZONE_COLORS.get(z, "#888") for z in zone_sum["zone"]],
                line=dict(color="#111C2D", width=3),
            ),
            textposition="outside",
            textinfo="percent+label",
            textfont=dict(family="Inter", size=11, color="#7A8FA8"),
        ))
        total_annotation = dict(
            text=f"<b>{int(zone_sum['closing'].sum()):,}</b><br>"
                 f"<span style='font-size:10px'>total</span>",
            x=0.5, y=0.5, font=dict(family="Rajdhani", size=22, color="#E8EDF5"),
            showarrow=False,
        )
        fig_donut.update_layout(
            **chart_layout(showlegend=False, height=360,
                           margin=dict(l=20, r=20, t=24, b=20)),
            annotations=[total_annotation],
        )
        st.plotly_chart(fig_donut, use_container_width=True, config={"displayModeBar": False})

# ── Charts row 2 ───────────────────────────────────────────────────────────────
col_c, col_d = st.columns([3, 2])

with col_c:
    section_heading("Office-wise Outstanding Paras")
    off_df = df_num.groupby(["office", "zone"])["closing"].sum().reset_index()
    off_order = (
        off_df.groupby("office")["closing"].sum()
        .sort_values(ascending=True).index
    )
    off_df["office"] = pd.Categorical(off_df["office"], categories=off_order, ordered=True)
    fig_off = px.bar(
        off_df.sort_values("office"),
        x="closing", y="office",
        color="zone", color_discrete_map=ZONE_COLORS, orientation="h",
        labels={"closing": "Outstanding Paras", "office": "", "zone": "Zone"},
    )
    fig_off.update_traces(marker_line_width=0)
    fig_off.update_layout(**chart_layout(xaxis_title="Outstanding Paras"))
    st.plotly_chart(fig_off, use_container_width=True, config={"displayModeBar": False})

with col_d:
    section_heading("Month-on-Month Trend")
    trend = (
        df[df["zone"].isin(selected_zones)].dropna(subset=["closing"])
        .groupby(["month", "month_dt", "zone"])["closing"].sum()
        .reset_index().sort_values("month_dt")
    )
    fig_trend = go.Figure()
    for zone in selected_zones:
        zd = trend[trend["zone"] == zone]
        if zd.empty:
            continue
        fig_trend.add_trace(go.Scatter(
            x=zd["month"], y=zd["closing"],
            name=ZONE_LABELS[zone],
            mode="lines+markers",
            line=dict(color=ZONE_COLORS[zone], width=2.5, shape="spline"),
            marker=dict(size=8, color=ZONE_COLORS[zone],
                        line=dict(width=2, color="#111C2D")),
        ))
    fig_trend.update_layout(**chart_layout(yaxis_title="Outstanding Paras"))
    st.plotly_chart(fig_trend, use_container_width=True, config={"displayModeBar": False})

# ── Zone performance analysis ──────────────────────────────────────────────────
st.markdown("<div style='height:0.4rem;'></div>", unsafe_allow_html=True)
section_heading("Zone Performance Analysis")

# ── Zone scorecards ────────────────────────────────────────────────────────────
zone_card_cols = st.columns(max(len(selected_zones), 1))
for i, zone in enumerate(selected_zones):
    zd = df_num[df_num["zone"] == zone]
    if zd.empty:
        continue
    closing_z   = int(zd["closing"].sum())
    opening_z   = int(zd["opening"].sum())
    raised_z    = int(zd["raised"].sum())
    settled_z   = int(zd["settled"].sum())
    exposure_z  = opening_z + raised_z
    s_rate_z    = round(settled_z / exposure_z * 100, 1) if exposure_z else 0.0

    if df_prev_all is not None:
        zd_prev  = df_prev_all[df_prev_all["zone"] == zone].dropna(subset=["closing"])
        prev_c_z = int(zd_prev["closing"].sum()) if not zd_prev.empty else None
        mom_z    = closing_z - prev_c_z if prev_c_z is not None else None
    else:
        mom_z = None

    color = ZONE_COLORS[zone]
    bar_w_z = f"{min(s_rate_z, 100):.1f}%"

    if mom_z is not None:
        mc = "#D94F3D" if mom_z > 0 else ("#38B089" if mom_z < 0 else "#5A7090")
        ma = "▲" if mom_z > 0 else ("▼" if mom_z < 0 else "—")
        mom_html_z = (
            f'<div style="font-family:Inter,sans-serif;font-size:0.63rem;color:{mc};margin-top:6px;">'
            f'{ma} {abs(mom_z):,} vs {prev_month}</div>'
        )
    else:
        mom_html_z = ""

    perf_label = "Efficient" if s_rate_z >= 4 else ("Moderate" if s_rate_z >= 1 else "Stagnant")
    perf_color = "#38B089" if s_rate_z >= 4 else ("#E8813A" if s_rate_z >= 1 else "#D94F3D")

    with zone_card_cols[i]:
        st.markdown(
            f'<div style="background:#12202F;border:1px solid {color}28;border-top:2px solid {color};'
            f'border-radius:4px;padding:16px 18px 14px;">'
            f'<div style="display:flex;justify-content:space-between;align-items:baseline;margin-bottom:6px;">'
            f'<span style="font-family:Rajdhani,sans-serif;font-size:1rem;font-weight:700;'
            f'letter-spacing:0.1em;color:{color};">{zone}</span>'
            f'<span style="font-family:Inter,sans-serif;font-size:0.58rem;font-weight:600;'
            f'letter-spacing:0.06em;color:{perf_color};background:{perf_color}18;'
            f'padding:2px 7px;border-radius:10px;">{perf_label}</span>'
            f'</div>'
            f'<div style="font-family:Inter,sans-serif;font-size:0.58rem;color:#5A7090;margin-bottom:2px;">'
            f'{ZONE_LABELS[zone]}</div>'
            f'<div style="font-family:Rajdhani,sans-serif;font-size:2.2rem;font-weight:700;'
            f'color:#E8EDF5;line-height:1;">{closing_z:,}</div>'
            f'<div style="font-family:Inter,sans-serif;font-size:0.58rem;color:#3A4E63;margin-bottom:10px;">'
            f'outstanding paras</div>'
            f'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:4px;">'
            f'<span style="font-family:Inter,sans-serif;font-size:0.6rem;color:#5A7090;">Settlement Rate</span>'
            f'<span style="font-family:Rajdhani,sans-serif;font-size:0.9rem;font-weight:600;color:{color};">'
            f'{s_rate_z:.1f}%</span>'
            f'</div>'
            f'<div style="background:#1A2A40;border-radius:2px;height:3px;">'
            f'<div style="width:{bar_w_z};height:3px;background:{color};border-radius:2px;"></div></div>'
            f'{mom_html_z}'
            f'</div>',
            unsafe_allow_html=True,
        )

st.markdown("<div style='height:1.2rem;'></div>", unsafe_allow_html=True)

# ── Settlement efficiency + pressure matrix ────────────────────────────────────
col_e, col_f = st.columns([3, 2])

with col_e:
    section_heading("Settlement Efficiency by Zone")

    perf_all = (
        df[df["zone"].isin(selected_zones)].dropna(subset=["closing"])
        .groupby(["zone", "month", "month_dt"])[["opening", "raised", "settled", "closing"]]
        .sum().reset_index().sort_values("month_dt")
    )
    perf_all["settlement_rate"] = perf_all.apply(
        lambda r: round(r["settled"] / (r["opening"] + r["raised"]) * 100, 2)
        if (r["opening"] + r["raised"]) > 0 else 0.0,
        axis=1,
    )
    perf_all["zone_label"] = perf_all["zone"].map(ZONE_LABELS)

    fig_eff = go.Figure()
    months_all = sorted(perf_all["month_dt"].unique())
    for mdt in months_all:
        md = perf_all[perf_all["month_dt"] == mdt].sort_values("zone")
        label = pd.Timestamp(mdt).strftime("%B %Y")
        is_selected = (label == selected_month)
        fig_eff.add_trace(go.Bar(
            name=label,
            x=md["zone_label"],
            y=md["settlement_rate"],
            marker=dict(
                color=[ZONE_COLORS.get(z, "#888") for z in md["zone"]],
                opacity=1.0 if is_selected else 0.35,
                line=dict(width=0),
            ),
            text=[f"{v:.1f}%" for v in md["settlement_rate"]],
            textposition="outside",
            textfont=dict(
                family="Rajdhani", size=12,
                color=["#E8EDF5" if is_selected else "#5A7090"] * len(md),
            ),
        ))

    avg_rate = perf_all[perf_all["month"] == selected_month]["settlement_rate"].mean()
    fig_eff.add_hline(
        y=avg_rate, line_dash="dot", line_color="#5A7090", line_width=1.5,
        annotation_text=f"Avg {avg_rate:.1f}%",
        annotation_font=dict(family="Inter", size=10, color="#5A7090"),
        annotation_position="top right",
    )
    fig_eff.update_layout(**chart_layout(
        barmode="group",
        yaxis_title="Settlement Rate (%)",
        yaxis_ticksuffix="%",
    ))
    st.plotly_chart(fig_eff, use_container_width=True, config={"displayModeBar": False})

with col_f:
    section_heading("Para Pressure Matrix")

    scatter_df = perf_all[perf_all["month"] == selected_month].copy()

    if not scatter_df.empty:
        # Quadrant lines at medians
        med_rate    = scatter_df["settlement_rate"].median()
        med_closing = scatter_df["closing"].median()

        fig_scatter = go.Figure()

        # Soft quadrant backgrounds
        x_max = scatter_df["settlement_rate"].max() * 1.6 + 1
        y_max = scatter_df["closing"].max() * 1.3

        fig_scatter.add_shape(type="rect", x0=0, x1=med_rate, y0=med_closing, y1=y_max,
                               fillcolor="rgba(217,79,61,0.04)", line_width=0)
        fig_scatter.add_shape(type="rect", x0=med_rate, x1=x_max, y0=0, y1=med_closing,
                               fillcolor="rgba(56,176,137,0.04)", line_width=0)
        fig_scatter.add_shape(type="line", x0=med_rate, x1=med_rate, y0=0, y1=y_max,
                               line=dict(color="#1A2A40", width=1, dash="dot"))
        fig_scatter.add_shape(type="line", x0=0, x1=x_max, y0=med_closing, y1=med_closing,
                               line=dict(color="#1A2A40", width=1, dash="dot"))

        for _, row in scatter_df.iterrows():
            z = row["zone"]
            color = ZONE_COLORS.get(z, "#888")
            bubble_size = max(16, min(52, int(row.get("raised", 0) / 8 + 18)))
            fig_scatter.add_trace(go.Scatter(
                x=[row["settlement_rate"]],
                y=[row["closing"]],
                mode="markers+text",
                marker=dict(size=bubble_size, color=color, opacity=0.88,
                            line=dict(width=2, color="#111C2D")),
                text=[z],
                textposition="middle center",
                textfont=dict(family="Rajdhani", size=11, color="#0A1628"),
                hovertemplate=(
                    f"<b>{ZONE_LABELS[z]}</b><br>"
                    f"Settlement Rate: {row['settlement_rate']:.1f}%<br>"
                    f"Outstanding: {int(row['closing']):,}<br>"
                    f"Raised: {int(row.get('raised', 0)):,}<extra></extra>"
                ),
                showlegend=False,
            ))

        fig_scatter.add_annotation(x=med_rate * 0.3, y=y_max * 0.92,
                                    text="High Burden · Low Efficiency",
                                    showarrow=False,
                                    font=dict(family="Inter", size=9, color="#D94F3D"),
                                    opacity=0.7)
        fig_scatter.add_annotation(x=x_max * 0.78, y=med_closing * 0.25,
                                    text="Low Burden · High Efficiency",
                                    showarrow=False,
                                    font=dict(family="Inter", size=9, color="#38B089"),
                                    opacity=0.7)

        fig_scatter.update_layout(**chart_layout(
            xaxis=dict(title="Settlement Rate (%)", ticksuffix="%",
                       gridcolor="#1A2A40", zerolinecolor="#1A2A40",
                       tickfont=dict(color="#7A8FA8", size=11), range=[0, x_max]),
            yaxis=dict(title="Outstanding Paras", gridcolor="#1A2A40",
                       zerolinecolor="#1A2A40", tickfont=dict(color="#7A8FA8", size=11),
                       range=[0, y_max]),
            showlegend=False, height=360,
        ))
        st.plotly_chart(fig_scatter, use_container_width=True, config={"displayModeBar": False})

# ── MoM net change ─────────────────────────────────────────────────────────────
if len(month_labels) > 1:
    st.markdown("<div style='height:0.5rem;'></div>", unsafe_allow_html=True)
    section_heading("Month-on-Month Net Change by Zone")

    mom_df = (
        df[df["zone"].isin(selected_zones)].dropna(subset=["closing", "opening"])
        .groupby(["zone", "month", "month_dt"])[["opening", "closing"]]
        .sum().reset_index().sort_values(["zone", "month_dt"])
    )
    mom_df["net_change"] = mom_df["closing"] - mom_df["opening"]
    mom_df["zone_label"] = mom_df["zone"].map(ZONE_LABELS)

    fig_mom = go.Figure()
    for _, row in mom_df.iterrows():
        nc = row["net_change"]
        bar_color = "#D94F3D" if nc > 0 else ("#38B089" if nc < 0 else "#5A7090")
        sign = "+" if nc > 0 else ""
        fig_mom.add_trace(go.Bar(
            x=[f"{row['zone_label']}<br><span style='font-size:10px'>{row['month']}</span>"],
            y=[nc],
            marker=dict(color=bar_color, opacity=0.85, line=dict(width=0)),
            text=[f"{sign}{nc:,}"],
            textposition="outside",
            textfont=dict(family="Rajdhani", size=12, color=bar_color),
            showlegend=False,
            hovertemplate=(
                f"<b>{row['zone_label']}</b> · {row['month']}<br>"
                f"Net Change: {sign}{nc:,}<br>"
                f"Opening: {int(row['opening']):,}<br>"
                f"Closing: {int(row['closing']):,}<extra></extra>"
            ),
        ))

    fig_mom.add_hline(y=0, line_color="#2B4069", line_width=1.5)
    fig_mom.update_layout(**chart_layout(
        yaxis_title="Net Change (Closing − Opening)",
        showlegend=False,
        height=300,
    ))
    st.markdown(
        '<div style="font-family:Inter,sans-serif;font-size:0.65rem;color:#5A7090;margin-bottom:4px;">'
        'Green bars = backlog reducing &nbsp;·&nbsp; Red bars = backlog growing</div>',
        unsafe_allow_html=True,
    )
    st.plotly_chart(fig_mom, use_container_width=True, config={"displayModeBar": False})

# ── Detail table ───────────────────────────────────────────────────────────────
st.markdown("<div style='height:0.4rem;'></div>", unsafe_allow_html=True)
section_heading("Paras Status — Office Detail")

tbl = (
    df_num.groupby(["zone", "office"])[["opening", "raised", "settled", "closing"]]
    .sum().reset_index()
)
tbl.columns = ["Zone", "Office", "Opening", "Raised", "Settled", "Closing"]
tbl["Net Change"] = tbl["Closing"] - tbl["Opening"]
tbl = tbl.sort_values(["Zone", "Closing"], ascending=[True, False])

st.dataframe(
    tbl, use_container_width=True, hide_index=True,
    column_config={
        "Opening": st.column_config.NumberColumn(format="%d"),
        "Raised": st.column_config.NumberColumn(format="%d"),
        "Settled": st.column_config.NumberColumn(format="%d"),
        "Closing": st.column_config.NumberColumn(format="%d"),
        "Net Change": st.column_config.NumberColumn(format="%+d"),
    },
)

# ── Zone totals ────────────────────────────────────────────────────────────────
st.markdown("<div style='height:1rem;'></div>", unsafe_allow_html=True)
section_heading("Zone Totals")

zone_totals = (
    tbl.groupby("Zone")[["Opening", "Raised", "Settled", "Closing", "Net Change"]]
    .sum().reset_index()
)
st.dataframe(
    zone_totals, use_container_width=True, hide_index=True,
    column_config={
        "Opening": st.column_config.NumberColumn(format="%d"),
        "Raised": st.column_config.NumberColumn(format="%d"),
        "Settled": st.column_config.NumberColumn(format="%d"),
        "Closing": st.column_config.NumberColumn(format="%d"),
        "Net Change": st.column_config.NumberColumn(format="%+d"),
    },
)

# ── Footer ─────────────────────────────────────────────────────────────────────
st.markdown(
    '<div style="height:1px;background:#1A2A40;margin:2rem 0 1rem;"></div>'
    '<div style="font-family:JetBrains Mono,monospace;font-size:0.58rem;color:#3A4E63;'
    'display:flex;justify-content:space-between;">'
    '<span>Source: Google Drive &nbsp;·&nbsp; IAW / Paras Status</span>'
    '<span>Refreshes every 5 min &nbsp;·&nbsp; MoHUA / IAW</span>'
    '</div>',
    unsafe_allow_html=True,
)
